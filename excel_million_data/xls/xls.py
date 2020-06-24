import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet, Cell
import csv
import abc
from datetime import datetime


class Column(metaclass=abc.ABCMeta):
    def __init__(self, cols: tuple):
        self._name = cols[0].value
        self._result = ()
        self.cols = cols[1:]

    @property
    def name(self):
        return self._name

    @property
    def result(self):
        return self._result

    def _parse(self, cols: tuple) -> tuple:
        pass

    @abc.abstractmethod
    def filter_btw(self, _from, _to) -> tuple:
        pass


class DateColumn(Column):
    def __init__(self, cols: tuple):
        super(DateColumn, self).__init__(cols)
        # this col should be as datetime type but being as a string
        self.cols = self._parse(self.cols)

    def _parse(self, cols: tuple):
        return tuple(datetime.strptime(c.value, '%Y-%m-%d') for i,c in enumerate(cols))

    def filter_btw(self, _from: datetime, _to: datetime) -> tuple:
        return tuple(
            filter(lambda c: _from <= c <= _to, self.cols)
        )


def read_xls(filename: str) -> Workbook:
    return openpyxl.load_workbook(filename)


def read_sheet(wb: Workbook, filename: str) -> Worksheet:
    return wb.get_sheet_by_name(filename)


def create_xls_from_csv(sheet_name: str, filename_from: str, filename_to: str):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    with open(filename_from) as f:
        reader = csv.reader(f, delimiter=',')
        for row in reader:
            ws.append(row)

    wb.save(filename_to)


def create_sheet(wb: Workbook, name: str) -> Worksheet:
    return wb.create_sheet(name)


def get_header(ws: Worksheet) -> list:
    return [col.value for col in next(ws.rows)[:ws.max_column]]


def crate_col_mapper(ws: Worksheet) -> dict:
    XLS_INDEX = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T',
                 'U', 'V', 'W', 'X', 'Y', 'Z']

    index = XLS_INDEX[:ws.max_column]
    header = get_header(ws)
    return {
        h: i for i, h in zip(index, header)
    }


# def parse(ws:Worksheet):
#     for row in ws.rows:
#         for c in enumerate(i,c)
#         tuple(datetime.strptime(c.value, '%Y-%m-%d') for i, c in enumerate(cols))
#



def filter_by_month(ws: Worksheet, col_name: str, crate_col_mapper: dict):
    # col_name is col name like 日付
    column = DateColumn(ws[crate_col_mapper[col_name]])
    rows = ws.rows

    cols = column.filter_btw(_from=datetime(year=2013, month=2, day=7), _to=datetime(year=2019, month=2, day=10))
    print(column.name)
    print(cols)


def pivot(filename):
    wb = read_xls(filename)
    print(wb.sheetnames)
    ws = wb.get_sheet_by_name("売上実績")
    p = ws._pivots
    print(p)
